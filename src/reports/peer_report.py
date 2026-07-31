from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.analytics.peer import PeerEngine


class PeerReport:

    def __init__(self):

        engine = PeerEngine()

        engine.calculate_percentiles()

        self.df = engine.df.copy()

    # -----------------------------------------------------

    def generate(self):

        wb = Workbook()

        wb.remove(wb.active)

        peer_groups = sorted(
            self.df["peer_group"]
            .dropna()
            .unique()
        )

        green = PatternFill(
            fill_type="solid",
            fgColor="90EE90"
        )

        yellow = PatternFill(
            fill_type="solid",
            fgColor="FFF59D"
        )

        red = PatternFill(
            fill_type="solid",
            fgColor="FFB6B6"
        )

        gold = PatternFill(
            fill_type="solid",
            fgColor="FFD966"
        )

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="4F81BD"
        )

        for group in peer_groups:

            sheet = wb.create_sheet(
                title=str(group)[:31]
            )

            temp = self.df[
                self.df["peer_group"] == group
            ].copy()

            temp = temp.sort_values(
                "composite_score_percentile",
                ascending=False
            )

            cols = [

                "company_name",

                "composite_score",

                "composite_score_percentile",

                "return_on_equity_pct",

                "return_on_equity_pct_percentile",

                "net_profit_margin_pct",

                "net_profit_margin_pct_percentile",

                "debt_to_equity",

                "debt_to_equity_percentile",

                "free_cash_flow_cr",

                "free_cash_flow_cr_percentile",

                "is_benchmark"

            ]

            temp = temp.reindex(columns=cols)

            # --------------------------
            # Header
            # --------------------------

            for c, col in enumerate(temp.columns, start=1):

                cell = sheet.cell(
                    row=1,
                    column=c
                )

                cell.value = col

                cell.font = Font(
                    bold=True,
                    color="FFFFFF"
                )

                cell.fill = header_fill

            # --------------------------
            # Data
            # --------------------------

            for r, row in enumerate(
                temp.itertuples(index=False),
                start=2
            ):

                for c, value in enumerate(
                    row,
                    start=1
                ):

                    cell = sheet.cell(
                        row=r,
                        column=c
                    )

                    cell.value = value

                # Benchmark Highlight
                if row[-1]:

                    for c in range(
                        1,
                        len(cols) + 1
                    ):

                        sheet.cell(
                            row=r,
                            column=c
                        ).fill = gold

                # Composite Percentile
                p = row[2]

                if p is not None:

                    if p >= 75:

                        sheet.cell(
                            row=r,
                            column=3
                        ).fill = green

                    elif p >= 25:

                        sheet.cell(
                            row=r,
                            column=3
                        ).fill = yellow

                    else:

                        sheet.cell(
                            row=r,
                            column=3
                        ).fill = red

            # --------------------------
            # Median Row
            # --------------------------

            median = temp.median(
                numeric_only=True
            )

            row_no = len(temp) + 3

            sheet.cell(
                row=row_no,
                column=1
            ).value = "MEDIAN"

            sheet.cell(
                row=row_no,
                column=1
            ).font = Font(
                bold=True
            )

            for c, col in enumerate(
                temp.columns[1:],
                start=2
            ):

                if col in median:

                    sheet.cell(
                        row=row_no,
                        column=c
                    ).value = round(
                        median[col],
                        2
                    )

            # --------------------------
            # Auto Width
            # --------------------------

            for column in sheet.columns:

                length = max(

                    len(str(cell.value))
                    if cell.value is not None
                    else 0

                    for cell in column

                )

                sheet.column_dimensions[
                    get_column_letter(
                        column[0].column
                    )
                ].width = min(
                    length + 3,
                    35
                )

        wb.save(
            "exports/peer_comparison.xlsx"
        )

        print()

        print("=" * 70)
        print("Peer Excel Report Generated")
        print("=" * 70)
        print("exports/peer_comparison.xlsx")